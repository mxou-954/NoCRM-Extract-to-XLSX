import os
import re
import html
import requests
from datetime import datetime
from os.path import join, dirname
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# Configuration
# ──────────────────────────────────────────────
dotenv_path = join(dirname(__file__), '.env')
load_dotenv(dotenv_path)

NOCRM_API_KEY = os.environ.get("NOCRM_API_KEY")
NOCRM_SUBDOMAIN = os.environ.get("NOCRM_SUBDOMAIN")

if not NOCRM_API_KEY or not NOCRM_SUBDOMAIN:
    raise ValueError("Variables d'environnement manquantes (NOCRM_API_KEY, NOCRM_SUBDOMAIN)")

BASE_URL = f"https://{NOCRM_SUBDOMAIN}.nocrm.io/api/v2"
HEADERS = {
    "X-API-KEY": NOCRM_API_KEY,
    "Content-Type": "application/json"
}

# ──────────────────────────────────────────────
# 1. Récupération des leads (avec pagination + filtre strict)
# ──────────────────────────────────────────────
def fetch_all_leads(step_id=None, limit=100, max_leads=None):
    """Récupère les leads avec filtre STRICT côté Python sur step_id."""
    all_leads = []
    offset = 0
    api_calls = 0

    while True:
        params = {"limit": limit, "offset": offset}
        if step_id:
            params["step_id"] = step_id

        response = requests.get(f"{BASE_URL}/leads", headers=HEADERS, params=params)
        api_calls += 1

        if response.status_code != 200:
            print(f"  ❌ Erreur API: {response.status_code} - {response.text}")
            break

        leads = response.json()
        if not leads:
            break

        # FILTRE STRICT : ne garder que les leads du bon step_id
        if step_id:
            before = len(leads)
            leads = [l for l in leads if l.get("step_id") == step_id]
            skipped = before - len(leads)
            if skipped:
                print(f"  ⚠️  {skipped} lead(s) hors étape ignoré(s)")

        all_leads.extend(leads)
        print(f"  → {len(all_leads)} leads valides (appel API #{api_calls})")

        if max_leads and len(all_leads) >= max_leads:
            all_leads = all_leads[:max_leads]
            break
        if len(response.json()) < limit:
            break
        offset += limit

    return all_leads


# ──────────────────────────────────────────────
# 2. Nettoyage description
# ──────────────────────────────────────────────
def clean_description(raw):
    """Nettoie la description (HTML ou texte brut)."""
    if not raw:
        return ""
    text = re.sub(r'<br\s*/?>|</p>|</div>|</li>', '\n', raw, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    text = html.unescape(text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


# ──────────────────────────────────────────────
# 3. Parsing description → champs entreprise
# ──────────────────────────────────────────────
def parse_description(description):
    clean = clean_description(description)
    if not clean:
        return {}

    fields = {}
    patterns = {
        "SIREN": r"SIREN\s*:\s*(.+)",
        "NAF": r"NAF\s*:\s*([^-–—\n]+)",
        "Effectif": r"Effectif\s*:\s*(.+)",
        "Adresse": r"Adresse\s*:\s*(.+)",
        "Chiffre d'affaires": r"Chiffre d[''']affaire[s]?[^:]*:\s*(.+)",
        "Résultat net": r"R[ée]sultat\s+net[^:]*:\s*(.+)",
        "Site web": r"Site\s+web\s*:\s*(.+)",
        "Budget transport": r"Budget\s+transport[^:]*:\s*(.+)",
        "Description": r"Description\s*:\s*(.+)",
    }

    for key, pattern in patterns.items():
        match = re.search(pattern, clean, re.IGNORECASE)
        if match:
            fields[key] = match.group(1).strip()

    return fields


# ──────────────────────────────────────────────
# 4. Parsing description → contacts
# ──────────────────────────────────────────────
def parse_contacts(description):
    clean = clean_description(description)
    if not clean:
        return []

    contacts = []
    blocks = re.split(r'\s*-{5,}\s*', clean)

    for block in blocks:
        nom = re.search(r"Nom\s*:\s*(.+)", block)
        if not nom:
            continue

        contact = {
            "Nom": nom.group(1).strip(),
            "Fonction": "",
            "Téléphone": "",
            "Email": "",
            "Source LinkedIn": ""
        }

        fonction = re.search(r"Fonction\s*:\s*(.+)", block)
        if fonction:
            contact["Fonction"] = fonction.group(1).strip()

        tel = re.search(r"T[ée]l[ée]phone\s*:\s*(.+)", block)
        if tel:
            contact["Téléphone"] = tel.group(1).strip()

        email_match = re.search(r"Email\s*:\s*(.+)", block)
        if email_match:
            contact["Email"] = email_match.group(1).strip()

        source = re.search(r"Source\s*:\s*(https?://\S+|.+)", block)
        if source:
            contact["Source LinkedIn"] = source.group(1).strip()

        contacts.append(contact)

    return contacts


# ──────────────────────────────────────────────
# 5. Extraction structurée d'un lead
# ──────────────────────────────────────────────
def extract_lead_data(lead):
    raw_desc = lead.get("description", "")
    desc_fields = parse_description(raw_desc)
    contacts = parse_contacts(raw_desc)

    ext = lead.get("extended_info", {}).get("fields", {})

    entreprise = {
        "ID Lead": lead.get("id", ""),
        "Titre": lead.get("title", ""),
        "Étape": lead.get("step", ""),
        "Tags": ", ".join(lead.get("tags", [])) if lead.get("tags") else "",
        "Créé le": lead.get("created_at", ""),
        "Mis à jour le": lead.get("updated_at", ""),
        "Status": lead.get("status", ""),
        "Amount": lead.get("amount", ""),
        "Prochaine action": lead.get("next_action_at", ""),
        "Date de rappel": lead.get("remind_date", ""),
        "Heure de rappel": lead.get("remind_time", ""),
        "Créé par": lead.get("created_from", ""),
        "SIREN": desc_fields.get("SIREN", ""),
        "NAF": desc_fields.get("NAF", ""),
        "Effectif": desc_fields.get("Effectif", ""),
        "Adresse": desc_fields.get("Adresse", ext.get("address", "")),
        "Chiffre d'affaires": desc_fields.get("Chiffre d'affaires", ""),
        "Résultat net": desc_fields.get("Résultat net", ""),
        "Site web": desc_fields.get("Site web", ext.get("web", "")),
        "Budget transport": desc_fields.get("Budget transport", ""),
        "Description activité": desc_fields.get("Description", ""),
    }

    return entreprise, contacts


# ──────────────────────────────────────────────
# 6. Export Excel
# ──────────────────────────────────────────────
def export_to_excel(leads_data, filename=None):
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_leads_nocrm_{timestamp}.xlsx"

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    cell_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(vertical='center', wrap_text=True)

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_data(ws, start_row=2):
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.font = cell_font
                cell.alignment = cell_align
                cell.border = thin_border

    def auto_width(ws, min_w=10, max_w=50):
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, min_w), max_w)

    # Feuille 1 : Entreprises
    ws_ent = wb.active
    ws_ent.title = "Entreprises"
    if leads_data:
        ent_headers = list(leads_data[0][0].keys())
        ws_ent.append(ent_headers)
        style_header(ws_ent)
        for entreprise, _ in leads_data:
            ws_ent.append(list(entreprise.values()))
        style_data(ws_ent)
        auto_width(ws_ent)
        ws_ent.auto_filter.ref = ws_ent.dimensions

    # Feuille 2 : Contacts
    ws_contacts = wb.create_sheet("Contacts")
    contact_headers = ["ID Lead", "Titre Lead", "Nom", "Fonction", "Téléphone", "Email", "Source LinkedIn"]
    ws_contacts.append(contact_headers)
    style_header(ws_contacts)
    for entreprise, contacts in leads_data:
        for c in contacts:
            ws_contacts.append([
                entreprise["ID Lead"],
                entreprise["Titre"],
                c.get("Nom", ""),
                c.get("Fonction", ""),
                c.get("Téléphone", ""),
                c.get("Email", ""),
                c.get("Source LinkedIn", ""),
            ])
    style_data(ws_contacts)
    auto_width(ws_contacts)
    ws_contacts.auto_filter.ref = ws_contacts.dimensions

    # Feuille 3 : Résumé
    ws_resume = wb.create_sheet("Résumé")
    ws_resume.append(["Métrique", "Valeur"])
    style_header(ws_resume)
    total_leads = len(leads_data)
    total_contacts = sum(len(contacts) for _, contacts in leads_data)
    ws_resume.append(["Total leads exportés", total_leads])
    ws_resume.append(["Total contacts extraits", total_contacts])
    ws_resume.append(["Date d'export", datetime.now().strftime("%d/%m/%Y %H:%M")])
    style_data(ws_resume)
    auto_width(ws_resume)

    wb.save(filename)
    return filename


# ──────────────────────────────────────────────
# 7. Main
# ──────────────────────────────────────────────
def main():
    print("=" * 50)
    print("  Export leads noCRM → Excel")
    print("=" * 50)

    # ⚠️ CONFIGURER ICI
    STEP_ID = 267810
    MAX_LEADS = None

    print(f"\nRécupération leads step_id={STEP_ID} (max={MAX_LEADS})...")
    leads = fetch_all_leads(step_id=STEP_ID, max_leads=MAX_LEADS)
    print(f"\n{len(leads)} leads 'Top prospects' récupérés")

    if not leads:
        print("Aucun lead trouvé.")
        return

    print("\nExtraction des données...")
    leads_data = []
    for lead in leads:
        entreprise, contacts = extract_lead_data(lead)
        leads_data.append((entreprise, contacts))
        print(f"  #{entreprise['ID Lead']} {entreprise['Titre']:30s} → {len(contacts)} contact(s)")

    total_contacts = sum(len(c) for _, c in leads_data)
    print(f"\n{len(leads_data)} entreprises, {total_contacts} contacts")

    print("\nGénération Excel...")
    filename = export_to_excel(leads_data)
    print(f"\nFichier exporté : {filename}")


if __name__ == "__main__":
    main()